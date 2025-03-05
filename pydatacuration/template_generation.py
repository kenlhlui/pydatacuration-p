"""Generates a report based on a template and data from a JSON file."""
from pathlib import Path

import jinja2
import orjson
import pandas as pd
from docxtpl import DocxTemplate
from jinja2 import Template
from openpyxl import load_workbook


RES_DIR = Path('res')


class GenerateReport:
    """Generates a report based on a template and data from a JSON file."""

    def __init__(self, workdir: Path) -> None:
        """Initializes the GenerateReport class.

        Args:
            workdir (Path): The working directory.
        """
        self.workdir = workdir

    @staticmethod
    def read_template_json() -> dict:
        """Reads the template.json file and returns it as a dictionary.

        Returns:
            dict: The template as a dictionary.
        """
        with RES_DIR.joinpath('template.json').open(encoding='utf-8') as file:
            return orjson.loads(file.read())

    def generate_report_xlsx(self, template_dict: dict) -> None:
        """Fill a formatted Excel template with data using Jinja2 for variable replacement.

        Args:
            template_dict (dict): Dictionary of data to inject into the template
        """
        # Load the workbook with formatting
        template_path = Path(RES_DIR, 'template_new.xlsx')
        workbook = load_workbook(template_path)
        sheet = workbook.active

        # Create Jinja2 environment
        env = jinja2.Environment()

        # Process each cell in the worksheet to replace Jinja2 variables
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '{%' in cell.value:
                    # This cell contains a Jinja2 template
                    template = env.from_string(cell.value)
                    try:
                        # Render the template with our data
                        cell.value = template.render(template_dict=template_dict)
                    except Exception as e:
                        print(f'Error rendering template in cell {cell.coordinate}: {e}')

        # Save the modified workbook
        excel_path_obj = self.workdir.joinpath('log_files', 'render_log_new.xlsx')
        workbook.save(excel_path_obj)
        print(f'\nReport generated. See the {str(excel_path_obj)} file for the spreadsheet report.')

    def generate_report_doc(self, template_dict: dict) -> None:
        """Generates a report based on the provided template dictionary.

        Args:
            template_dict (dict): The template dictionary.
        """
        # Load the template
        template_path = Path(RES_DIR, 'template_new.docx')
        doc = DocxTemplate(template_path)

        # Render the template with the template_dict
        doc.render({'template_dict': template_dict})  # TEMP fix. Need to restructure the template_dict

        # Save the rendered document
        doc_path = self.workdir.joinpath('log_files', 'render_log.docx')
        doc.save(doc_path)
        print(f'\nReport generated. See the {str(doc_path)} file for the docx report.')
