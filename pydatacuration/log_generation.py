"""Generates a report based on a template and data from a JSON file."""
import sys
from datetime import datetime
from datetime import timezone
from pathlib import Path

import jinja2
import jmespath
import orjson
import yaml
from docxtpl import DocxTemplate
from markitdown import MarkItDown
from openpyxl import load_workbook

from .custom_logging import CustomLogger


RES_DIR = Path('res')


class GenerateLog:
    """Generates a report based on a template and data from a JSON file."""

    def __init__(self, log_dir: Path, base_url: str, ds_metadata: dict, ticket_number: str) -> None:
        """Initializes the GenerateLog class.

        Args:
            workdir (Path): The working directory.
            base_url (str): The base URL of the repository.
            ds_metadata (dict): The dataset metadata.
            ticket_number (str): The ticket number.
        """
        self.log_dir = log_dir
        self.timestamp = self._get_timestamp()
        self.ds_metadata = ds_metadata
        self.base_url = base_url
        self.dataset_info_dict = self._get_dataset_info()
        self.ticket_number = ticket_number
        self.logger = CustomLogger.get_logger(__name__)

    @staticmethod
    def _get_timestamp() -> str:
        """Returns the current timestamp and return it as a dictionary.

        Returns:
            str : The current timestamp in 'YYYY-MM-DD HH:MM:SS' format.
        """
        utc_now = datetime.now(tz=timezone.utc)
        local_now = utc_now.astimezone()
        return local_now.strftime('%Y-%m-%d %H:%M:%S')

    @staticmethod
    def _convert_to_markdown(doc_path: Path) -> Path:
        """Converts the report to markdown format.

        Args:
            doc_path (Path): The path to the document to be converted.

        """
        md = MarkItDown()
        result = md.convert(doc_path)
        # Save the markdown file
        md_path = doc_path.with_suffix('.md')
        if result.text_content:
            with md_path.open('w', encoding='utf-8') as file:
                file.write(result.text_content)

        return md_path

    @staticmethod
    def read_template_json() -> dict:
        """Reads the template.json file and returns it as a dictionary.

        Returns:
            dict: The template as a dictionary.
        """
        with RES_DIR.joinpath('template.json').open(encoding='utf-8') as file:
            return orjson.loads(file.read())

    def _get_dataset_info(self) -> dict:
        """Returns the dataset information as a dictionary.

        Returns:
            dict: The dataset information.
        """
        search_string = """{
        DatasetTitle: join(', ', data.latestVersion.metadataBlocks.citation.fields[?typeName==`title`].value),
        DatasetPersistentId: data.latestVersion.datasetPersistentId,
        ID: data.latestVersion.id}"""
        dataset_info_dict = jmespath.search(search_string, self.ds_metadata)
        # Add 'DatasetURL' by parsing the base_url and the DatasetPersistentId
        dataset_info_dict['DatasetURL'] = f'{self.base_url}/dataset.xhtml?persistentId={dataset_info_dict.get('DatasetPersistentId', None)}'  # noqa: E501

        return dataset_info_dict

    def _get_config_info(self) -> dict:
        """Reads the config.yaml file and returns it as a dictionary.

        Returns:
            dict: The config as a dictionary.
        """
        with RES_DIR.joinpath('config.yaml').open(encoding='utf-8') as file:
            config_dict = yaml.safe_load(file)
            # Add the ticket number to the config dictionary
            config_dict['ticket_number'] = self.ticket_number
            return config_dict

    def generate_report_xlsx(self, template_dict: dict) -> None:
        """Fill a formatted Excel template with data using Jinja2 for variable replacement.

        Args:
            template_dict (dict): Dictionary of data to inject into the template
        """
        # Load the workbook with formatting
        template_path = Path(RES_DIR, 'template_new.xlsx')
        workbook = load_workbook(template_path)
        sheet = workbook.active

        # Create Jinja2 environment
        env = jinja2.Environment()

        # Process each cell in the worksheet to replace Jinja2 variables
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '{%' in cell.value:
                    # This cell contains a Jinja2 template
                    template = env.from_string(cell.value)
                    try:
                        # Render the template with our data
                        cell.value = template.render(template_dict=template_dict)
                    except Exception as e:
                        self.logger.error(f'Error rendering template in cell {cell.coordinate}: {e}')

        # Save the modified workbook
        excel_path_obj = self.log_dir.joinpath('render_log_new.xlsx')
        workbook.save(excel_path_obj)
        self.logger.print(f'Excel Spreadsheet curation log saved at: {str(excel_path_obj)}')

    def generate_report_doc(self, template_dict: dict, level: str) -> None:
        """Generates a report based on the provided template dictionary.

        Args:
            template_dict (dict): The template dictionary.
            level (str): The level of the report.
        """
        # Load the template
        if level == 'medium':
            template_path = Path(RES_DIR, 'template_medium.docx')
        elif level == 'high':
            template_path = Path(RES_DIR, 'template_high.docx')
        else:
            self.logger.error(f'Invalid level: {level}. Must be "medium" or "high".')
            sys.exit(1)

        doc = DocxTemplate(template_path)

        # Render the document with the provided context
        doc.render({'template_dict': template_dict,  # TEMP fix. Need to restructure the template_dict
                    'timestamp': self.timestamp,
                    'dataset': self.dataset_info_dict,
                    'project_info': self._get_config_info()})

        # Save the rendered document
        doc_path = self.log_dir.joinpath(f'log_{level}-level.docx')
        doc.save(doc_path)
        self.logger.print(f'{level.upper()}-level Word curation log saved at: {str(doc_path)}')

        # Convert the report to markdown format
        md_path = self._convert_to_markdown(doc_path)
        self.logger.print(f'Converted {level.upper()}-level Word curation log to Markdown format at: {str(md_path)}')

    def generate_project_metadata(self) -> None:
        """Generates project metadata (project_info) to JSON file."""
        meta_path = self.log_dir.joinpath('project_info.json')
        with meta_path.open('w', encoding='utf-8') as file:
            file.write(orjson.dumps(self._get_config_info(), option=orjson.OPT_INDENT_2).decode('utf-8'))
            self.logger.print(f'Project metadata saved at: {str(meta_path)}')
